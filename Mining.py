import json
import pandas as pd
from mlxtend.frequent_patterns import apriori
from mlxtend.frequent_patterns import fpgrowth
from mlxtend.frequent_patterns import association_rules
import matplotlib as mpl
from matplotlib import pyplot as plt
from matplotlib import colors as mcolors
from matplotlib import _color_data as mcd
import string
import math
import numpy as np
from tqdm import tqdm
import copy
import math
import openpyxl
from openpyxl.styles.borders import Border, Side
import os
import re

SAVE_PATH = os.getcwd()

class Mining:

    def __init__(self):
        self.new_column_list = []
        self.unique_category_III_list = []
        self.present_in_all_samples = []
        self.not_present_in_any_samples = []
        self.lower_than_abundance_cutoff_value_in_all_samples = []
        self.to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples = []
        self.to_be_dropped_as_present_in_all_samples = []
        self.support_value_list_equal = []
        self.support_value_list_less_or_equal = []
        self.unique_support_values = []
        self.closed_itemset = [] #it will store a frozen set
        self.closed_itemsets = []
        self.closed_itemset_support = []
        self.closed_itemset_supports = []
        self.closed_itemset_assoc_support = []
        self.closed_itemset_assoc_supports = []
        self.closed_itemset_antecedents = []
        self.closed_itemset_consequents = []
        self.closed_itemset_antecedent_supports = []
        self.closed_itemset_consequent_supports = []
        self.closed_itemset_confidences = []
        self.closed_itemset_lifts = []
        self.closed_itemset_leverages = []
        self.closed_itemset_convictions = []
        self.closed_itemset_zhangs_metrics = []
        self.maximal_frequent_itemsets = []
        self.maximal_frequent_itemset_supports = []
        self.maximal_frequent_itemset_assoc_supports = []
        self.maximal_frequent_itemset_antecedents = []
        self.maximal_frequent_itemset_consequents = []
        self.maximal_frequent_itemset_antecedent_supports = []
        self.maximal_frequent_itemset_consequent_supports = []
        self.maximal_frequent_itemset_confidences = []
        self.maximal_frequent_itemset_lifts = []
        self.maximal_frequent_itemset_leverages = []
        self.maximal_frequent_itemset_convictions = []
        self.maximal_frequent_itemset_zhangs_metrics = []
        self.evaluated_set = []
        self.used_taxa_list = []
        self.count_number_list = []
        self.frozen_itemset_length_list_final = []   
        self.frequent_itemset_sample_number_list_final = []
        self.closed_itemset_sample_number_list_final = []
        self.maximal_frequent_itemset_sample_number_list_final = []
        self.general_sample_number_list_final = []
        self.frequent_association_rules_list_final = []
        
        self.input_dataframe = pd.DataFrame()
        self.func_dataframe_category_value = pd.DataFrame()
        self.mining_output_df = pd.DataFrame()
        self.frequent_itemsets = pd.DataFrame()
        self.present_in_all_samples_df = pd.DataFrame(columns = ['present_in_all_samples'])
        self.not_present_in_any_samples_df = pd.DataFrame(columns = ['not_present_in_any_samples'])
        self.lower_than_abundance_cutoff_value_in_all_samples_df = pd.DataFrame(columns = ['abundance_value_is_lower_than_cutoff_value_in_all_samples'])
        self.to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples_df = \
	pd.DataFrame(columns = ['to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples'])
        self.to_be_dropped_as_present_in_all_samples_df = pd.DataFrame(columns = ['to_be_dropped_as_present_in_all_samples'])
        self.frequent_itemset_df = pd.DataFrame(columns = ['frequent_itemsets', 'origin_sample', 'support', 'length'])
        self.frequent_itemset_summary_stat_df = pd.DataFrame()
        self.frequent_itemset_df_by_item = pd.DataFrame(columns = ['item', 'itemset_length_list', 'origin_sample', 'count'])
        self.closed_itemset_df = pd.DataFrame(columns = ['closed_itemsets', 'origin_sample', 'support', 'length'])
        self.closed_itemset_summary_stat_df = pd.DataFrame()
        self.closed_itemset_df_by_item = pd.DataFrame(columns = ['item', 'itemset_length_list', 'origin_sample', 'count'])
        self.maximum_frequent_itemset_df = pd.DataFrame(columns = ['maximum_frequent_itemsets', 'origin_sample', 'support', 'length'])
        self.maximum_frequent_itemset_summary_stat_df = pd.DataFrame()
        self.maximum_frequent_itemset_df_by_item = pd.DataFrame(columns = ['item', 'itemset_length_list', 'origin_sample', 'count'])     
        self.association_dataframe = pd.DataFrame()
        self.association_dataframe_summary_stat_df = pd.DataFrame()
        self.closed_association_rules_dataframe  = pd.DataFrame(columns = ['closed_antecedents', 'closed_consequents','length_antecedents','length_consequents','origin_sample_antecedents',
									   'origin_sample_consequents', 'antecedent support', 'consequent support', 'support', 'confidence', 'lift', 'leverage',
									   'conviction', 'zhangs_metric'])
        self.closed_association_rules_stat_df = pd.DataFrame()
        self.max_frequent_association_rules_dataframe  = pd.DataFrame(columns = ['maximum_frequent_antecedents', 'maximum_frequent_consequents','length_antecedents','length_consequents',
										 'origin_sample_antecedents', 'origin_sample_consequents', 'antecedent support', 'consequent support', 'support',
										 'confidence', 'lift', 'leverage', 'conviction', 'zhangs_metric'])
        self.max_frequent_association_rules_stat_df = pd.DataFrame()
        self.closed_itemset_df_by_sample = pd.DataFrame()
        self.closed_itemset_association_df_by_sample = pd.DataFrame()
        self.maximum_frequent_itemset_df_by_sample = pd.DataFrame()
        self.maximum_frequent_itemset_association_df_by_sample = pd.DataFrame()
        self.closed_itemset_df_sample_matrix = pd.DataFrame()
        self.closed_itemset_association_df_sample_matrix = pd.DataFrame()
        self.maximum_frequent_itemset_df_sample_matrix = pd.DataFrame()
        self.maximum_frequent_itemset_association_df_sample_matrix = pd.DataFrame()        
        self.unique_support_values_dictionary_equal = {}
        self.unique_support_values_dictionary_less_or_equal = {}      
        self.itemset_category = ""
        self.count_number = int(0) 
        
    def set_input_dataframe(self, assign):
        self.input_dataframe = assign

    def get_input_dataframe(self):
        return self.input_dataframe
    
    def set_present_in_all_samples(self, assign):
        self.present_in_all_samples = assign

    def set_not_present_in_any_samples(self, assign):
        self.not_present_in_any_samples = assign

    def set_lower_than_abundance_cutoff_value_in_all_samples(self, assign):
        self.lower_than_abundance_cutoff_value_in_all_samples = assign

    def set_to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples(self, assign):
        self.to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples = assign

    def set_to_be_dropped_as_present_in_all_samples(self, assign):
        self.to_be_dropped_as_present_in_all_samples = assign

    def set_present_in_all_samples_df(self, assign):
        self.present_in_all_samples_df = assign

    def set_not_present_in_any_samples_df(self, assign):
        self.not_present_in_any_samples_df = assign

    def set_lower_than_abundance_cutoff_value_in_all_samples_df(self, assign):
        self.lower_than_abundance_cutoff_value_in_all_samples_df = assign

    def set_to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples_df(self, assign):
        self.to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples_df = assign

    def set_to_be_dropped_as_present_in_all_samples_df(self, assign):
        self.to_be_dropped_as_present_in_all_samples_df = assign

    def get_unique_support_values_dictionary_equal(self):
        return self.unique_support_values_dictionary_equal

    def get_unique_support_values_dictionary_less_or_equal(self):
        return self.unique_support_values_dictionary_less_or_equal

    def set_support_value_list_equal(self, assign):
        self.support_value_list_equal = assign

    def get_support_value_list_equal(self):
        return self.support_value_list_equal

    def set_support_value_list_less_or_equal(self, assign):
        self.support_value_list_less_or_equal = assign

    def get_support_value_list_less_or_equal(self):
        return self.support_value_list_less_or_equal

    def set_unique_support_values(self, assign):
        self.unique_support_values = assign

    def get_unique_support_values(self):
        return self.unique_support_values

    def set_closed_itemsets(self, assign):
        self.closed_itemsets = assign
    
    def get_closed_itemsets(self):
        return self.closed_itemsets

    def set_closed_itemset_supports(self, assign):
        self.closed_itemset_supports = assign
    
    def get_closed_itemset_supports(self):
        return self.closed_itemset_supports 

    def set_maximal_frequent_itemsets(self, assign):
        self.maximal_frequent_itemsets = assign

    def get_maximal_frequent_itemsets(self):
        return self.maximal_frequent_itemsets

    def set_frequent_itemsets(self, assign):
        self.frequent_itemsets = assign

    def get_frequent_itemsets(self):
        return self.frequent_itemsets

    def set_itemset_category(self, assign):
        self.itemset_category = assign

    def get_itemset_category(self):
        return self.itemset_category

    def set_closed_itemset(self, assign):
        self.closed_itemset = assign

    def get_closed_itemset(self):
        return self.closed_itemset

    def set_closed_itemset_support(self, assign):
        self.closed_itemset_support = assign

    def get_closed_itemset_support(self):
        return self.closed_itemset_support

    def set_evaluated_set(self, assign):
        self.evaluated_set = assign

    def get_evaluated_set(self):
        return self.evaluated_set

    def set_frequent_itemset_df(self, assign):
        self.frequent_itemset_df = assign

    def get_frequent_itemset_df(self):
        return self.frequent_itemset_df

    def set_closed_itemset_df(self, assign):
        self.closed_itemset_df = assign

    def get_closed_itemset_df(self):
        return self.closed_itemset_df

    def set_maximum_frequent_itemset_df(self, assign):
        self.maximum_frequent_itemset_df = assign

    def get_maximum_frequent_itemset_df(self):
        return self.maximum_frequent_itemset_df

    def set_frequent_itemset_summary_stat_df(self, assign):
        self.frequent_itemset_summary_stat_df = assign

    def get_frequent_itemset_summary_stat_df(self):
        return self.frequent_itemset_summary_stat_df

    def set_closed_itemset_summary_stat_df(self, assign):
        self.closed_itemset_summary_stat_df = assign

    def get_closed_itemset_summary_stat_df(self):
        return self.closed_itemset_summary_stat_df

    def set_maximum_frequent_itemset_summary_stat_df(self, assign):
        self.maximum_frequent_itemset_summary_stat_df = assign

    def get_maximum_frequent_itemset_summary_stat_df(self):
        return self.maximum_frequent_itemset_summary_stat_df

    def set_frequent_itemset_df_by_item(self, assign):
        self.frequent_itemset_df_by_item = assign

    def get_frequent_itemset_df_by_item(self):
        return self.frequent_itemset_df_by_item

    def set_closed_itemset_df_by_item(self, assign):
        self.closed_itemset_df_by_item = assign

    def get_closed_itemset_df_by_item(self):
        return self.closed_itemset_df_by_item

    def set_maximum_frequent_itemset_df_by_item(self, assign):
        self.maximum_frequent_itemset_df_by_item = assign

    def get_maximum_frequent_itemset_df_by_item(self):
        return self.maximum_frequent_itemset_df_by_item

    def set_used_taxa_list(self, assign):
        self.used_taxa_list = assign

    def get_used_taxa_list(self):
        return self.used_taxa_list

    def set_count_number_list(self, assign):
        self.count_number_list = assign

    def get_count_number_list(self):
        return self.count_number_list

    def set_frozen_itemset_length_list_final(self, assign):
        self.frozen_itemset_length_list_final = assign

    def get_frozen_itemset_length_list_final(self):
        return self.frozen_itemset_length_list_final

    def set_count_number(self, assign):
        self.count_number = assign

    def get_count_number(self):
        return self.count_number

    def set_mining_output_df(self, assign):
        self.mining_output_df = assign

    def get_mining_output_df(self):
        return self.mining_output_df

    def set_frequent_itemset_sample_number_list_final(self, assign):
        self.frequent_itemset_sample_number_list_final = assign

    def get_frequent_itemset_sample_number_list_final(self):
        return self.frequent_itemset_sample_number_list_final

    def set_closed_itemset_sample_number_list_final(self, assign):
        self.closed_itemset_sample_number_list_final = assign

    def get_closed_itemset_sample_number_list_final(self):
        return self.closed_itemset_sample_number_list_final

    def set_maximal_frequent_itemset_sample_number_list_final(self, assign):
        self.maximal_frequent_itemset_sample_number_list_final = assign

    def get_maximal_frequent_itemset_sample_number_list_final(self):
        return self.maximal_frequent_itemset_sample_number_list_final

    def set_general_sample_number_list_final(self, assign):
        self.general_sample_number_list_final  = assign

    def get_general_sample_number_list_final(self):
        return self.general_sample_number_list_final

    def set_association_dataframe(self, assign):
        self.association_dataframe = assign

    def get_association_dataframe(self):
        return self.association_dataframe

    def set_frequent_association_rules_list_final(self, assign):
        self.frequent_association_rules_list_final = assign

    def get_frequent_association_rules_list_final(self):
        return self.frequent_association_rules_list_final

    def set_closed_association_rules_dataframe(self, assign):
        self.closed_association_rules_dataframe = assign

    def get_closed_association_rules_dataframe(self):
        return self.closed_association_rules_dataframe

    def set_max_frequent_association_rules_dataframe(self, assign):
        self.max_frequent_association_rules_dataframe = assign   

    def get_max_frequent_association_rules_dataframe(self):
        return self.max_frequent_association_rules_dataframe

    def set_closed_itemset_antecedent(self, assign):
        self.closed_itemset_antecedent = assign

    def get_closed_itemset_antecedent(self):
        return self.closed_itemset_antecedent

    def set_closed_itemset_antecedents(self, assign):
        self.closed_itemset_antecedents = assign

    def get_closed_itemset_antecedents(self):
        return self.closed_itemset_antecedents

    def set_closed_itemset_consequent(self, assign):
        self.closed_itemset_consequent = assign

    def get_closed_itemset_consequent(self):
        return self.closed_itemset_consequent

    def set_closed_itemset_consequents(self, assign):
        self.closed_itemset_consequents = assign

    def get_closed_itemset_consequents(self):
        return self.closed_itemset_consequents

    def set_closed_itemset_assoc_support(self, assign):
        self.closed_itemset_assoc_support = assign

    def get_closed_itemset_assoc_support(self):
        return self.closed_itemset_assoc_support

    def set_closed_itemset_assoc_supports(self, assign):
        self.closed_itemset_assoc_supports = assign

    def get_closed_itemset_assoc_supports(self):
        return self.closed_itemset_assoc_supports

    def set_closed_itemset_antecedent_supports(self, assign):
        self.closed_itemset_antecedent_supports = assign

    def get_closed_itemset_antecedent_supports(self):
        return self.closed_itemset_antecedent_supports

    def set_closed_itemset_consequent_supports(self, assign):
        self.closed_itemset_consequent_supports = assign

    def get_closed_itemset_consequent_supports(self):
        return self.closed_itemset_consequent_supports 

    def set_closed_itemset_confidences(self, assign):
        self.closed_itemset_confidences = assign

    def get_closed_itemset_confidences(self):
        return self.closed_itemset_confidences

    def set_closed_itemset_lifts(self, assign):
        self.closed_itemset_lifts = assign

    def get_closed_itemset_lifts(self):
        return self.closed_itemset_lifts

    def set_closed_itemset_leverages(self, assign):
        self.closed_itemset_leverages = assign

    def get_closed_itemset_leverages(self):
        return self.closed_itemset_leverages

    def set_closed_itemset_convictions(self, assign):
        self.closed_itemset_convictions = assign

    def get_closed_itemset_convictions(self):
        return self.closed_itemset_convictions

    def set_closed_itemset_zhangs_metrics(self, assign):
        self.closed_itemset_zhangs_metrics = assign

    def get_closed_itemset_zhangs_metrics(self):
        return self.closed_itemset_zhangs_metrics

    def set_maximal_frequent_itemset_antecedents(self, assign):
        self.maximal_frequent_itemset_antecedents = assign

    def get_maximal_frequent_itemset_antecedents(self):
        return self.maximal_frequent_itemset_antecedents

    def set_maximal_frequent_itemset_consequents(self, assign):
        self.maximal_frequent_itemset_consequents = assign

    def get_maximal_frequent_itemset_consequents(self):
        return self.maximal_frequent_itemset_consequents

    def set_maximal_frequent_itemset_supports(self, assign):
        self.maximal_frequent_itemset_supports = assign

    def get_maximal_frequent_itemset_supports(self):
        return self.maximal_frequent_itemset_supports

    def set_maximal_frequent_itemset_assoc_supports(self, assign):
        self.maximal_frequent_itemset_assoc_supports = assign

    def get_maximal_frequent_itemset_assoc_supports(self):
        return self.maximal_frequent_itemset_assoc_supports

    def set_maximal_frequent_itemset_antecedent_supports(self, assign):
        self.maximal_frequent_itemset_antecedent_supports = assign

    def get_maximal_frequent_itemset_antecedent_supports(self):
        return self.maximal_frequent_itemset_antecedent_supports 

    def set_maximal_frequent_itemset_consequent_supports(self, assign):
        self.maximal_frequent_itemset_consequent_supports = assign

    def get_maximal_frequent_itemset_consequent_supports(self):
        return self.maximal_frequent_itemset_consequent_supports

    def set_maximal_frequent_itemset_confidences(self, assign):
        self.maximal_frequent_itemset_confidences = assign

    def get_maximal_frequent_itemset_confidences(self):
        return self.maximal_frequent_itemset_confidences

    def set_maximal_frequent_itemset_lifts(self, assign):
        self.maximal_frequent_itemset_lifts = assign

    def get_maximal_frequent_itemset_lifts(self):
        return self.maximal_frequent_itemset_lifts

    def set_maximal_frequent_itemset_leverages(self, assign):
        self.maximal_frequent_itemset_leverages = assign

    def get_maximal_frequent_itemset_leverages(self):
        return self.maximal_frequent_itemset_leverages

    def set_maximal_frequent_itemset_convictions(self, assign):
        self.maximal_frequent_itemset_convictions = assign

    def get_maximal_frequent_itemset_convictions(self):
        return self.maximal_frequent_itemset_convictions

    def set_maximal_frequent_itemset_zhangs_metrics(self, assign):
        self.maximal_frequent_itemset_zhangs_metrics = assign

    def get_maximal_frequent_itemset_zhangs_metrics(self):
        return self.maximal_frequent_itemset_zhangs_metrics

    def set_association_dataframe_summary_stat_df(self, assign):
        self.association_dataframe_summary_stat_df = assign

    def get_association_dataframe_summary_stat_df(self):
        return self.association_dataframe_summary_stat_df

    def set_closed_association_rules_stat_df(self, assign):
        self.closed_association_rules_stat_df = assign

    def get_closed_association_rules_stat_df(self):
        return self.closed_association_rules_stat_df

    def set_max_frequent_association_rules_stat_df(self, assign):
        self.max_frequent_association_rules_stat_df = assign

    def get_max_frequent_association_rules_stat_df(self):
        return self.max_frequent_association_rules_stat_df

    def set_closed_itemset_df_by_sample(self, assign):
        self.closed_itemset_df_by_sample = assign

    def get_closed_itemset_df_by_sample(self):
        return self.closed_itemset_df_by_sample

    def set_closed_itemset_association_df_by_sample(self, assign):
        self.closed_itemset_association_df_by_sample = assign

    def get_closed_itemset_association_df_by_sample(self):
        return self.closed_itemset_association_df_by_sample

    def set_maximum_frequent_itemset_df_by_sample(self, assign):
        self.maximum_frequent_itemset_df_by_sample = assign

    def get_maximum_frequent_itemset_df_by_sample(self):
        return self.maximum_frequent_itemset_df_by_sample

    def set_maximum_frequent_itemset_association_df_by_sample(self, assign):
        self.maximum_frequent_itemset_association_df_by_sample = assign

    def get_maximum_frequent_itemset_association_df_by_sample(self):
        return self.maximum_frequent_itemset_association_df_by_sample

    def set_closed_itemset_df_sample_matrix(self, assign):
        self.closed_itemset_df_sample_matrix = assign

    def get_closed_itemset_df_sample_matrix(self):
        return self.closed_itemset_df_sample_matrix

    def set_closed_itemset_association_df_sample_matrix(self, assign):
        self.closed_itemset_association_df_sample_matrix = assign

    def get_closed_itemset_association_df_sample_matrix(self):
        return self.closed_itemset_association_df_sample_matrix

    def set_maximum_frequent_itemset_df_sample_matrix(self, assign):
        self.maximum_frequent_itemset_df_sample_matrix = assign

    def get_maximum_frequent_itemset_df_sample_matrix(self):
        return self.maximum_frequent_itemset_df_sample_matrix

    def set_maximum_frequent_itemset_association_df_sample_matrix(self, assign):
        self.maximum_frequent_itemset_association_df_sample_matrix = assign

    def get_maximum_frequent_itemset_association_df_sample_matrix(self):
        return self.maximum_frequent_itemset_association_df_sample_matrix

        
    '''
    Update the dataframe for the further steps.
    '''
    def update_dataframes(self, input_file, update_functionality_ID_file = ""):
        self.output_dataframe = pd.DataFrame()
        try:
            self.input_dataframe = pd.read_csv(f'{SAVE_PATH}/{input_file}.csv', low_memory = False)
        except:
            print('The input file is missing.')
        try:
            self.input_dataframe = self.input_dataframe.loc[:, ~self.input_dataframe.columns.str.contains('^Unnamed')]
        except:
            pass
        self.input_dataframe_original = copy.deepcopy(self.input_dataframe)
        self.input_dataframe = self.input_dataframe.iloc[:,5:]
        self.update_columns(self.input_dataframe, input_file, self.output_dataframe, update_functionality_ID_file)

        self.output_dataframe = pd.DataFrame()
        
        self.unique_category_III_list = self.input_dataframe_original["Category III"].unique()
        
        for i, item in enumerate(self.unique_category_III_list):
            self.input_dataframe_category_value = self.input_dataframe_original[self.input_dataframe_original["Category III"] == self.unique_category_III_list[i]]
            self.input_dataframe_category_value = self.input_dataframe_category_value.iloc[:,5:]
            self.update_columns(self.input_dataframe_category_value, f"{input_file}_{item}", self.output_dataframe, update_functionality_ID_file)

            
    '''
    Replace functionality ID-s with Functionality names
    '''
    def update_functionality_IDs(self, input_dataframe, update_functionality_ID_file, output_dataframe):
        self.functionality_origin = ""
        output_dataframe = copy.deepcopy(input_dataframe)
        
        try:
            self.functionality_ID_file_dataframe = pd.read_csv(f'{SAVE_PATH}/{update_functionality_ID_file}.csv', low_memory = False)
        except:
            print('The input file is missing.')

        if str(input_dataframe.columns[0])[0:3] == "GO:":
            self.functionality_origin = "GO:"
        elif str(input_dataframe.columns[0])[0:4] == "IPR0":
            self.functionality_origin = "IPRO:"
        
        for i, item in enumerate(output_dataframe.columns.values):
            for j, thing in enumerate(self.functionality_ID_file_dataframe['id']):
                if item == thing:
                    self.new_name = self.functionality_origin + self.functionality_ID_file_dataframe.at[j, 'attributes.description'] + '-' + self.functionality_ID_file_dataframe.at[j, 'attributes.lineage']
                    self.new_name = self.new_name.replace(' ', '_')
                    output_dataframe.columns.values[i] = self.new_name
        
    
    '''
    Update column titles to ensure that certain characters are removed from the titles.
    '''
    def update_columns(self, input_dataframe, input_file, output_dataframe, update_functionality_ID_file = ""):
        self.new_column_list = []
        output_dataframe = copy.deepcopy(input_dataframe)
        
        if (str(input_dataframe.columns[0])[0:3] != "GO:") and (str(input_dataframe.columns[0])[0:4] != "IPR0"):
            for i, item in enumerate(input_dataframe.columns):
                self.cleaned_item = str(item).replace("(", "")
                self.cleaned_item = self.cleaned_item.replace(")", "")
                self.cleaned_item = self.cleaned_item.replace("'", "")
                self.cleaned_item = self.cleaned_item.replace("_", "+")
                self.cleaned_item = self.cleaned_item.replace(":", "_")
                self.split_item = self.cleaned_item.split(" ")
                self.new_column_list.append(self.split_item[0] + "_" + self.split_item[1])
            output_dataframe.columns = self.new_column_list
            output_dataframe.to_csv(f"{SAVE_PATH}//{input_file}_new_columns.csv")
            
        elif (str(input_dataframe.columns[0])[0:3] == "GO:") or (str(input_dataframe.columns[0])[0:4] == "IPR0"):
            if update_functionality_ID_file != "":
                self.update_functionality_IDs(input_dataframe, update_functionality_ID_file, output_dataframe) 
                output_dataframe.to_csv(f"{SAVE_PATH}//{input_file}_new_columns.csv")
                
            else:
                print("Columns were not updated as update_functionality_ID_file was not given.")
                output_dataframe.to_csv(f"{SAVE_PATH}//{input_file}_new_columns.csv")
            
        self.new_column_list = []
        
        return output_dataframe

    '''
    Export the results of exploratory data analysis
    '''
    def export_exploratory_data_evaluation(self, input_dataframe, output_name, table_position_1, table_position_2):
        if not os.path.exists(f"{SAVE_PATH}/{output_name}"):
            workbook = openpyxl.Workbook()
            workbook.create_sheet("Evaluation")
            workbook.save(f"{SAVE_PATH}/{output_name}")
        
        with pd.ExcelWriter(f"{SAVE_PATH}/{output_name}", engine="openpyxl", mode = 'a', if_sheet_exists = 'overlay') as writer:
            input_dataframe.to_excel(writer, sheet_name="Evaluation", startrow=table_position_1, startcol=table_position_2, header=True, index=True)
            workbook = writer.book
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
        
    '''
    Apply the selected threshold value on the transaction dataset
    '''
    def update_values(self, abundance_threshold, input_file):
        try:
            self.input_dataframe = pd.read_csv(f'{SAVE_PATH}/{input_file}.csv', low_memory = False, index_col = 0)
        except:
            print('The input file is missing.')
        try:
            self.input_dataframe = self.input_dataframe.loc[:, ~self.input_dataframe.columns.str.contains('^Unnamed')]
        except:
            pass
        if abundance_threshold != 0:
            for item in self.input_dataframe.columns:
                self.input_dataframe.loc[self.input_dataframe[f"{item}"] < abundance_threshold, f"{item}"] = False
                self.input_dataframe.loc[self.input_dataframe[f"{item}"] >= abundance_threshold, f"{item}"] = True
        self.input_dataframe.to_csv(f"{SAVE_PATH}//{input_file}_abund_cutoff.csv")
        
        for column in self.input_dataframe.columns:
            if len(self.input_dataframe[f"{column}"].unique()) == 1:
                if self.input_dataframe[f"{column}"].unique()[0] == True:
                    self.present_in_all_samples.append(column)
                if self.input_dataframe[f"{column}"].unique()[0] == False:
                    self.lower_than_abundance_cutoff_value_in_all_samples.append(column)
                if math.isnan(self.input_dataframe[f"{column}"].unique()[0]) == True:
                    self.not_present_in_any_samples.append(column)

        if len(self.present_in_all_samples) != 0:
            print(f"After the abundance cutoff value is applied, the following taxon/functionality is still present in all samples:\n{self.present_in_all_samples}\n")
            self.present_in_all_samples_df['present_in_all_samples'] = self.present_in_all_samples
            self.export_exploratory_data_evaluation(self.present_in_all_samples_df, f"{input_file}_presence_absence_evaluation.xlsx", table_position_1 = 0, table_position_2 = 1)

        if len(self.not_present_in_any_samples) != 0:
            print(f"Originally, the following taxon/functionality is NOT present in any samples:\n{self.not_present_in_any_samples}\n")
            self.not_present_in_any_samples_df['not_present_in_any_samples'] = self.not_present_in_any_samples
            self.export_exploratory_data_evaluation(self.not_present_in_any_samples_df, f"{input_file}_presence_absence_evaluation.xlsx", table_position_1 = 0, table_position_2 = 4)

        if len(self.lower_than_abundance_cutoff_value_in_all_samples) != 0:
            print(f"After the abundance cutoff value is applied, the following taxon/functionality is NOT present in any samples:\n{self.lower_than_abundance_cutoff_value_in_all_samples}\n")
            self.lower_than_abundance_cutoff_value_in_all_samples_df['abundance_value_is_lower_than_cutoff_value_in_all_samples'] = self.lower_than_abundance_cutoff_value_in_all_samples
            self.export_exploratory_data_evaluation(self.lower_than_abundance_cutoff_value_in_all_samples_df, f"{input_file}_presence_absence_evaluation.xlsx", table_position_1 = 0, table_position_2 = 7)               

        for item in self.input_dataframe.columns:
            self.input_dataframe.loc[self.input_dataframe[f"{item}"] != True, f"{item}"] = False
        self.input_dataframe.to_csv(f"{SAVE_PATH}//{input_file}_abund_filtered.csv")

        for column in self.input_dataframe.columns:
            if len(self.input_dataframe[f"{column}"].unique()) == 1:
                if self.input_dataframe[f"{column}"].unique()[0] == True:
                    self.to_be_dropped_as_present_in_all_samples.append(column)       
                if self.input_dataframe[f"{column}"].unique()[0] == False:
                    self.to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples.append(column)

        if len(self.to_be_dropped_as_present_in_all_samples) != 0:
            print(f"The following taxon/functionality is removed as those are present in all samples:\n{self.to_be_dropped_as_present_in_all_samples}\n")
            self.to_be_dropped_as_present_in_all_samples_df['to_be_dropped_as_present_in_all_samples'] = self.to_be_dropped_as_present_in_all_samples
            self.export_exploratory_data_evaluation(self.to_be_dropped_as_present_in_all_samples_df, f"{input_file}_presence_absence_evaluation.xlsx", table_position_1 = 0, table_position_2 = 10)

        if len(self.to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples) != 0:
            print(f"The following taxon/functionality is removed as those are NOT present in any samples:\n{self.to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples}\n")
            self.to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples_df['to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples'] =\
	    self.to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples
            self.export_exploratory_data_evaluation(self.to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples_df, f"{input_file}_presence_absence_evaluation.xlsx",
						    table_position_1 = 0, table_position_2 = 13)
            
        self.non_unique_columns = self.input_dataframe.nunique()

        if len(self.non_unique_columns) != 0:
            self.dropped_columns = self.non_unique_columns[self.non_unique_columns == 1].index
            try:
            	self.input_dataframe.drop(self.dropped_columns, axis = 1, inplace = True)
            except:
                print("The column titles are not present, so the columns can't be dropped")
        self.input_dataframe = self.input_dataframe.convert_dtypes()
        self.input_dataframe.astype('bool').dtypes

        self.set_present_in_all_samples([])
        self.set_not_present_in_any_samples([])
        self.set_lower_than_abundance_cutoff_value_in_all_samples([])
        self.set_to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples([])
        self.set_to_be_dropped_as_present_in_all_samples([])

        self.set_present_in_all_samples_df(pd.DataFrame(columns = ['present_in_all_samples']))
        self.set_not_present_in_any_samples_df(pd.DataFrame(columns = ['not_present_in_any_samples']))
        self.set_lower_than_abundance_cutoff_value_in_all_samples_df(pd.DataFrame(columns = ['abundance_value_is_lower_than_cutoff_value_in_all_samples']))
        self.set_to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples_df(pd.DataFrame(columns =\
	['to_be_dropped_as_not_present_in_any_samples_or_lower_than_abundance_cutoff_value_in_all_samples']))
        self.set_to_be_dropped_as_present_in_all_samples_df(pd.DataFrame(columns = ['to_be_dropped_as_present_in_all_samples']))

        return self.input_dataframe

    '''
    Calculate closed itemsets from frequent itemsets
    '''
    def calculate_closed_itemsets(self, input_dataframe,  input_dataframe_metric_column,  input_dataframe_itemset_column):

        self.set_unique_support_values(input_dataframe[f'{input_dataframe_metric_column}'].unique())

        for i in range(len(self.get_unique_support_values())):
            self.set_support_value_list_equal(list(input_dataframe.loc[input_dataframe[f'{ input_dataframe_metric_column}'] == self.get_unique_support_values()[i]][f'{input_dataframe_itemset_column}']))
            self.get_unique_support_values_dictionary_equal()[self.get_unique_support_values()[i]] = self.get_support_value_list_equal()


        for i, item in input_dataframe.iterrows():
            self.set_itemset_category("Closed")
            self.set_closed_itemset(item[f'{input_dataframe_itemset_column}'])
            self.set_closed_itemset_support(item[f'{input_dataframe_metric_column}'])
            self.set_evaluated_set(self.get_unique_support_values_dictionary_equal()[self.get_closed_itemset_support()])
            for thing in self.get_evaluated_set():
                if self.get_closed_itemset() != thing:
                    if(frozenset.issubset(self.get_closed_itemset(), thing)):
                        self.set_itemset_category("Not closed")
                        break
            if self.get_itemset_category() == "Closed":
                self.get_closed_itemsets().append(item[f'{input_dataframe_itemset_column}'])
                self.get_closed_itemset_supports().append(item[f'{input_dataframe_metric_column}'])

    '''
    Calculates closed itemsets for assocaiation rules
    '''
    def calculate_closed_itemsets_for_association_rules(self, input_dataframe,  input_dataframe_metric_column,  input_dataframe_itemset_column):

        self.set_unique_support_values(input_dataframe[f'{input_dataframe_metric_column}'].unique())
        
        for i in range(len(self.get_unique_support_values())):
            self.set_support_value_list_equal(list(input_dataframe.loc[input_dataframe[f'{ input_dataframe_metric_column}'] == self.get_unique_support_values()[i]][f'{input_dataframe_itemset_column}']))
            self.get_unique_support_values_dictionary_equal()[self.get_unique_support_values()[i]] = self.get_support_value_list_equal()
        
        
        for i, item in input_dataframe.iterrows():
            self.set_itemset_category("Closed")
            self.set_closed_itemset(item[f'{input_dataframe_itemset_column}'])
            self.set_closed_itemset_assoc_support(item[f'{input_dataframe_metric_column}'])
        
            self.set_evaluated_set(self.get_unique_support_values_dictionary_equal()[self.get_closed_itemset_assoc_support()])
            for thing in self.get_evaluated_set():
                if self.get_closed_itemset() != thing:
                    if(frozenset.issubset(self.get_closed_itemset(), thing)):
                        self.set_itemset_category("Not closed")
                        break
            if self.get_itemset_category() == "Closed":
                self.get_closed_itemsets().append(item[f'{input_dataframe_itemset_column}'])
                self.get_closed_itemset_assoc_supports().append(item[f'{input_dataframe_metric_column}'])
                self.get_closed_itemset_antecedents().append(item['antecedents'])
                self.get_closed_itemset_consequents().append(item['consequents'])
                self.get_closed_itemset_supports().append(item['support'])
                self.get_closed_itemset_antecedent_supports().append(item['antecedent support'])
                self.get_closed_itemset_consequent_supports().append(item['consequent support'])
                self.get_closed_itemset_confidences().append(item['confidence'])
                self.get_closed_itemset_lifts().append(item['lift'])
                self.get_closed_itemset_leverages().append(item['leverage'])
                self.get_closed_itemset_convictions().append(item['conviction'])
                self.get_closed_itemset_zhangs_metrics().append(item['zhangs_metric'])
        
    '''
    Calculate max frequent itemsets from frequent itemsets
    '''
    def calculate_max_frequent_itemsets(self, input_dataframe, input_dataframe_metric_column,  input_dataframe_itemset_column):

        self.set_unique_support_values(input_dataframe[f'{input_dataframe_metric_column}'].unique())

        for i in range(len(self.get_unique_support_values())):
            self.set_support_value_list_less_or_equal(list(input_dataframe.loc[input_dataframe[f'{input_dataframe_metric_column}'] <= self.get_unique_support_values()[i]][f'{input_dataframe_itemset_column}']))
            self.get_unique_support_values_dictionary_less_or_equal()[self.get_unique_support_values()[i]] = self.get_support_value_list_less_or_equal()

        for i, item in input_dataframe.iterrows():
            self.set_itemset_category("Closed")
            self.set_closed_itemset(item[f'{input_dataframe_itemset_column}'])
            self.set_closed_itemset_support(item[f'{input_dataframe_metric_column}'])
            self.set_evaluated_set(self.get_unique_support_values_dictionary_less_or_equal()[self.get_closed_itemset_support()])
            for thing in self.get_evaluated_set():
                if self.get_closed_itemset() != thing:
                     if(frozenset.issubset(self.get_closed_itemset(), thing)):
                         self.set_itemset_category("Not closed")
                         break
            if self.get_itemset_category() == "Closed":
                self.get_maximal_frequent_itemsets().append(item[f'{input_dataframe_itemset_column}'])
                self.get_maximal_frequent_itemset_supports().append(item[f'{input_dataframe_metric_column}'])

    '''
    Calculate max frequent itemsets for association rules
    '''
    def calculate_max_frequent_itemsets_for_association_rules(self, input_dataframe, input_dataframe_metric_column,  input_dataframe_itemset_column):

        self.set_unique_support_values(input_dataframe[f'{input_dataframe_metric_column}'].unique())
        
        for i in range(len(self.get_unique_support_values())):
            self.set_support_value_list_less_or_equal(list(input_dataframe.loc[input_dataframe[f'{ input_dataframe_metric_column}'] <= self.get_unique_support_values()[i]][f'{input_dataframe_itemset_column}']))
            self.get_unique_support_values_dictionary_less_or_equal()[self.get_unique_support_values()[i]] = self.get_support_value_list_less_or_equal()
            
                
        for i, item in input_dataframe.iterrows():
            self.set_itemset_category("Closed")
            self.set_closed_itemset(item[f'{input_dataframe_itemset_column}'])
            self.set_closed_itemset_assoc_support(item[f'{input_dataframe_metric_column}'])
            self.set_evaluated_set(self.get_unique_support_values_dictionary_less_or_equal()[self.get_closed_itemset_assoc_support()])
            for thing in self.get_evaluated_set():
                if self.get_closed_itemset() != thing:
                    if(frozenset.issubset(self.get_closed_itemset(), thing)):
                        self.set_itemset_category("Not closed")
                        break
            if self.get_itemset_category() == "Closed":
                self.get_maximal_frequent_itemsets().append(item[f'{input_dataframe_itemset_column}'])
                self.get_maximal_frequent_itemset_assoc_supports().append(item[f'{input_dataframe_metric_column}'])
                self.get_maximal_frequent_itemset_antecedents().append(item['antecedents'])
                self.get_maximal_frequent_itemset_consequents().append(item['consequents'])
                self.get_maximal_frequent_itemset_supports().append(item['support'])
                self.get_maximal_frequent_itemset_antecedent_supports().append(item['antecedent support'])
                self.get_maximal_frequent_itemset_consequent_supports().append(item['consequent support'])
                self.get_maximal_frequent_itemset_confidences().append(item['confidence'])
                self.get_maximal_frequent_itemset_lifts().append(item['lift'])
                self.get_maximal_frequent_itemset_leverages().append(item['leverage'])
                self.get_maximal_frequent_itemset_convictions().append(item['conviction'])
                self.get_maximal_frequent_itemset_zhangs_metrics().append(item['zhangs_metric'])

    '''
    Set lists used for closed and max frequent itemset calculation to zero
    '''
    def set_closed_and_max_frequent_itemset_lists_to_zero(self):
        self.set_closed_itemsets([])
        self.set_closed_itemset_supports([])
        self.set_maximal_frequent_itemsets([])
        self.set_maximal_frequent_itemset_supports([])
        self.set_frequent_itemset_sample_number_list_final([])
        self.set_closed_itemset_sample_number_list_final([])
        self.set_maximal_frequent_itemset_sample_number_list_final([])

        self.set_closed_itemset_assoc_support([])
        self.set_closed_itemset_assoc_supports([])
        self.set_closed_itemset_antecedent_supports([])
        self.set_closed_itemset_consequent_supports([])
        self.set_closed_itemset_confidences([])
        self.set_closed_itemset_lifts([])
        self.set_closed_itemset_leverages([])
        self.set_closed_itemset_convictions([])
        self.set_closed_itemset_zhangs_metrics([])
        self.set_closed_itemset_antecedents([])
        self.set_closed_itemset_consequents([])
        self.set_maximal_frequent_itemset_supports([])
        self.set_maximal_frequent_itemset_assoc_supports([])
        self.set_maximal_frequent_itemset_antecedents([])
        self.set_maximal_frequent_itemset_consequents([])     
        self.set_maximal_frequent_itemset_antecedent_supports([])
        self.set_maximal_frequent_itemset_consequent_supports([])
        self.set_maximal_frequent_itemset_confidences([])
        self.set_maximal_frequent_itemset_lifts([])
        self.set_maximal_frequent_itemset_leverages([])
        self.set_maximal_frequent_itemset_convictions([])
        self.set_maximal_frequent_itemset_zhangs_metrics([])



    
    '''
    Creates a dataframe sorted by item for FPGrowth and Apriori result evaluation
    '''
    def create_by_item_dataframe(self, itemset_dataframe, itemset_column, output_dataframe, input_dataframe):
        for i, item in enumerate(itemset_dataframe[f'{itemset_column}']):
            for thing in item:
                if thing not in self.get_used_taxa_list():
                    self.get_used_taxa_list().append(thing)
                    self.set_count_number(0)    
                    self.temp_list = []
                    for element in itemset_dataframe[f'{itemset_column}']:
                        for member in element:
                            if member == thing:
                                self.set_count_number(int(self.get_count_number()) + 1)
                                self.temp_list.append(len(element))
                    self.get_frozen_itemset_length_list_final().append(self.temp_list)
                    self.get_count_number_list().append(self.get_count_number())


        output_dataframe['item'] = self.get_used_taxa_list()
        output_dataframe['itemset_length_list'] = self.get_frozen_itemset_length_list_final()
        output_dataframe['count'] = self.get_count_number_list()


        for item in output_dataframe['item']:
            self.temp_list = []
            for element in input_dataframe.columns:
                if element == item:
                    for i, member in enumerate(input_dataframe[f"{element}"]):
                        if member == True:
                            self.temp_list.append(input_dataframe.index.values[i])
            self.get_general_sample_number_list_final().append(self.temp_list)

        output_dataframe['origin_sample'] = self.get_general_sample_number_list_final()
        output_dataframe.sort_values(by = 'count', ascending = False, inplace = True)
        output_dataframe.reset_index(drop = True, inplace = True)

        self.set_used_taxa_list([])
        self.set_count_number_list([])
        self.set_frozen_itemset_length_list_final([])
        self.set_general_sample_number_list_final([])
        

    '''
    Creates a dataframe sorted by sample for FPGrowth, Apriori result evaluation
    '''
    def create_by_sample_dataframe(self, itemset_dataframe, itemset_column, sample_column, output_dataframe):
        self.bool_outcome_list = []
        self.sample_number_list = []
        self.itemset_list = []
        
        for i, item in enumerate(itemset_dataframe[f'{sample_column}']):
            for number in item:
                for thing in number:
                    for member in itemset_dataframe[f'{sample_column}']:
                        for submember in member:
                            if thing in submember:
                                self.bool_outcome_list.append(True)
                    if False not in self.bool_outcome_list:
                        self.sample_number_list.append(thing)
                        self.itemset_list.append(itemset_dataframe.at[i, f'{itemset_column}'])
                    self.bool_outcome_list = []

        output_dataframe['sample_number'] = self.sample_number_list
        output_dataframe['itemsets'] = self.itemset_list
        output_dataframe.reset_index(drop = True, inplace = True)
                
        self.sample_number_list = []
        self.itemset_list = []


    '''
    Creates a dataframe sorted by sample for association rules result evaluation
    '''
    def create_by_sample_dataframe_for_association_rules(self, itemset_dataframe, itemset_list_antecedents, itemset_list_consequents, output_dataframe):
        self.bool_outcome_list = []
        self.sample_number_list = []
        self.itemset_list_antecedents = []
        self.itemset_list_consequents = []
        self.temp_itemset_list_antecedents_with_consequents = []
        self.itemset_list_antecedents_with_consequents = []
        
        for i, item in enumerate(itemset_dataframe['origin_sample_antecedents']):
            for number in item:
                for thing in number:
                    for member in itemset_dataframe['origin_sample_antecedents']:
                        for submember in member:
                            if thing in submember:
                                self.bool_outcome_list.append(True)
                    if False not in self.bool_outcome_list:
                        self.sample_number_list.append(thing)
                        self.itemset_list_antecedents.append(itemset_dataframe.at[i, f'{itemset_list_antecedents}'])
                        self.itemset_list_consequents.append(itemset_dataframe.at[i, f'{itemset_list_consequents}'])
                    self.bool_outcome_list = []
                    
        for i, item in enumerate(self.itemset_list_antecedents):
            self.unfrozen_set = [] 
            for member in item:
                self.unfrozen_set.append('#ANT_' + member)
            self.unfrozen_set.append(str('#CON_' + list(self.itemset_list_consequents[i])[0]))
            self.itemset_list_antecedents_with_consequents.append(frozenset(self.unfrozen_set))
                
        output_dataframe['common_sample_number'] = self.sample_number_list
        output_dataframe[f'{itemset_list_antecedents}'] = self.itemset_list_antecedents
        output_dataframe[f'{itemset_list_consequents}'] = self.itemset_list_consequents
        output_dataframe['antecedent_with_consequents'] = self.itemset_list_antecedents_with_consequents


        for i, item in enumerate(output_dataframe['common_sample_number']):
            if item not in self.sample_number_list:
                output_dataframe.drop(i, inplace = True)
                
        output_dataframe.reset_index(drop = True, inplace = True)
 

        self.sample_number_list = []
        self.itemset_list_antecedents = []
        self.itemset_list_consequents = []
        self.itemset_list_antecedents_with_consequents = []
        
    
    '''
    Assign lists to dataframes for frequent itemsets
    '''
    def assign_lists_to_dataframes(self):

        self.get_frequent_itemset_df()['frequent_itemsets'] = self.get_frequent_itemsets()['itemsets']
        self.get_frequent_itemset_df()['support'] = self.get_frequent_itemsets()['support']
        self.get_frequent_itemset_df()['length'] = self.get_frequent_itemset_df()['frequent_itemsets'].apply(lambda x: len(x))
        self.get_sample_origin(self.get_frequent_itemset_df(), self.get_mining_output_df(), 'frequent_itemsets', self.get_frequent_itemset_sample_number_list_final())
        self.get_frequent_itemset_df()['origin_sample'] = self.get_frequent_itemset_sample_number_list_final()
        self.get_frequent_itemset_df().sort_values(by = ['length', 'support'], ascending = [False, False], inplace = True)
        self.get_frequent_itemset_df().reset_index(drop = True, inplace = True)
        self.set_frequent_itemset_summary_stat_df(self.get_frequent_itemset_df()['length'].value_counts())
        self.create_by_item_dataframe(self.get_frequent_itemset_df(), 'frequent_itemsets', self.get_frequent_itemset_df_by_item(), self.get_mining_output_df())
        
        self.get_closed_itemset_df()['closed_itemsets'] = self.get_closed_itemsets()
        self.get_closed_itemset_df()['support'] = self.get_closed_itemset_supports()
        self.get_closed_itemset_df()['length'] = self.get_closed_itemset_df()['closed_itemsets'].apply(lambda x: len(x))
        self.get_sample_origin(self.get_closed_itemset_df(), self.get_mining_output_df(), 'closed_itemsets', self.get_closed_itemset_sample_number_list_final())
        self.get_closed_itemset_df()['origin_sample'] = self.get_closed_itemset_sample_number_list_final()
        self.get_closed_itemset_df().sort_values(by = ['length', 'support'], ascending = [False, False], inplace = True)
        self.get_closed_itemset_df().reset_index(drop = True, inplace = True)
        self.set_closed_itemset_summary_stat_df(self.get_closed_itemset_df()['length'].value_counts())
        self.create_by_item_dataframe(self.get_closed_itemset_df(), 'closed_itemsets', self.get_closed_itemset_df_by_item(), self.get_mining_output_df())
        self.create_by_sample_dataframe(self.get_closed_itemset_df(), 'closed_itemsets', 'origin_sample',  self.get_closed_itemset_df_by_sample())
        self.get_closed_itemset_df_by_sample().drop_duplicates(subset=['itemsets', 'sample_number'], inplace = True)
        self.get_closed_itemset_df_by_sample().sort_values(by = 'sample_number', inplace = True)
        self.get_closed_itemset_df_by_sample().reset_index(drop = True, inplace = True)
        self.set_closed_itemset_df_sample_matrix(self.create_itemset_mining_matrix(self.get_closed_itemset_df_by_sample(), 'sample_number', 'itemsets', self.get_closed_itemset_df_sample_matrix()))

        self.get_maximum_frequent_itemset_df()['maximum_frequent_itemsets'] = self.get_maximal_frequent_itemsets()
        self.get_maximum_frequent_itemset_df()['support'] = self.get_maximal_frequent_itemset_supports()
        self.get_maximum_frequent_itemset_df()['length'] = self.get_maximum_frequent_itemset_df()['maximum_frequent_itemsets'].apply(lambda x: len(x))
        self.get_sample_origin(self.get_maximum_frequent_itemset_df(), self.get_mining_output_df(), 'maximum_frequent_itemsets', self.get_maximal_frequent_itemset_sample_number_list_final())
        self.get_maximum_frequent_itemset_df()['origin_sample'] = self.get_maximal_frequent_itemset_sample_number_list_final()
        self.get_maximum_frequent_itemset_df().sort_values(by = ['length', 'support'], ascending = [False, False], inplace = True)
        self.get_maximum_frequent_itemset_df().reset_index(drop = True, inplace = True)
        self.set_maximum_frequent_itemset_summary_stat_df(self.get_maximum_frequent_itemset_df()['length'].value_counts())
        self.create_by_item_dataframe(self.get_maximum_frequent_itemset_df(), 'maximum_frequent_itemsets', self.get_maximum_frequent_itemset_df_by_item(), self.get_mining_output_df())
        self.create_by_sample_dataframe(self.get_maximum_frequent_itemset_df(), 'maximum_frequent_itemsets', 'origin_sample',  self.get_maximum_frequent_itemset_df_by_sample())
        self.get_maximum_frequent_itemset_df_by_sample().drop_duplicates(subset=['itemsets', 'sample_number'], inplace = True)
        self.get_maximum_frequent_itemset_df_by_sample().sort_values(by = 'sample_number', inplace = True)
        self.get_maximum_frequent_itemset_df_by_sample().reset_index(drop = True, inplace = True)
        self.set_maximum_frequent_itemset_df_sample_matrix(self.create_itemset_mining_matrix(self.get_maximum_frequent_itemset_df_by_sample(), 'sample_number', 'itemsets', self.get_maximum_frequent_itemset_df_sample_matrix()))
    
    '''
    Assign lists to dataframes for association rules
    '''
    def assign_lists_to_dataframes_for_association_rules(self, processed_data_origin_file):
        
        self.get_closed_association_rules_dataframe()['closed_antecedents'] = self.get_closed_itemset_antecedents()
        self.get_closed_association_rules_dataframe()['closed_consequents'] = self.get_closed_itemset_consequents()
        self.get_closed_association_rules_dataframe()['length_antecedents'] = self.get_closed_association_rules_dataframe()['closed_antecedents'].apply(lambda x: len(x))
        self.get_closed_association_rules_dataframe()['length_consequents'] = self.get_closed_association_rules_dataframe()['closed_consequents'].apply(lambda x: len(x))
        self.get_sample_origin_for_association_rules(self.get_closed_association_rules_dataframe(), 'closed_antecedents', processed_data_origin_file, self.get_closed_itemset_sample_number_list_final())
        self.get_closed_association_rules_dataframe()['origin_sample_antecedents'] = self.get_closed_itemset_sample_number_list_final()
        self.set_closed_itemset_sample_number_list_final([])
        self.get_sample_origin_for_association_rules(self.get_closed_association_rules_dataframe(), 'closed_consequents', processed_data_origin_file, self.get_closed_itemset_sample_number_list_final())
        self.get_closed_association_rules_dataframe()['origin_sample_consequents'] = self.get_closed_itemset_sample_number_list_final()
        self.set_closed_itemset_sample_number_list_final([])
        self.get_closed_association_rules_dataframe()['support'] = self.get_closed_itemset_supports()
        self.get_closed_association_rules_dataframe()['antecedent support'] = self.get_closed_itemset_antecedent_supports()
        self.get_closed_association_rules_dataframe()['consequent support'] = self.get_closed_itemset_consequent_supports()
        self.get_closed_association_rules_dataframe()['confidence'] = self.get_closed_itemset_confidences()
        self.get_closed_association_rules_dataframe()['lift'] = self.get_closed_itemset_lifts()
        self.get_closed_association_rules_dataframe()['leverage'] = self.get_closed_itemset_leverages()
        self.get_closed_association_rules_dataframe()['conviction'] = self.get_closed_itemset_convictions()
        self.get_closed_association_rules_dataframe()['zhangs_metric'] = self.get_closed_itemset_zhangs_metrics()
        self.get_closed_association_rules_dataframe().sort_values(by = ['length_antecedents', 'length_consequents', 'support'], ascending = [False, False, False], inplace = True)
        self.get_closed_association_rules_dataframe().reset_index(drop = True, inplace = True)
        self.set_closed_association_rules_stat_df(self.get_closed_association_rules_dataframe().value_counts(subset = ['length_antecedents', 'length_consequents']))
        self.create_by_sample_dataframe_for_association_rules(self.get_closed_association_rules_dataframe(), 'closed_antecedents', 'closed_consequents',  self.get_closed_itemset_association_df_by_sample())
        self.get_closed_itemset_association_df_by_sample().sort_values(by = 'common_sample_number', inplace = True)
        self.get_closed_itemset_association_df_by_sample().reset_index(drop = True, inplace = True)
        self.set_closed_itemset_association_df_sample_matrix(self.create_association_mining_matrix(self.get_closed_itemset_association_df_by_sample(), 'common_sample_number', 'antecedent_with_consequents',
												   self.get_closed_itemset_association_df_sample_matrix()))
    
        self.get_max_frequent_association_rules_dataframe()['maximum_frequent_antecedents'] = self.get_maximal_frequent_itemset_antecedents()
        self.get_max_frequent_association_rules_dataframe()['maximum_frequent_consequents'] = self.get_maximal_frequent_itemset_consequents()
        self.get_max_frequent_association_rules_dataframe()['length_antecedents'] = self.get_max_frequent_association_rules_dataframe()['maximum_frequent_antecedents'].apply(lambda x: len(x))
        self.get_max_frequent_association_rules_dataframe()['length_consequents'] = self.get_max_frequent_association_rules_dataframe()['maximum_frequent_consequents'].apply(lambda x: len(x))
        self.get_sample_origin_for_association_rules(self.get_max_frequent_association_rules_dataframe(), 'maximum_frequent_antecedents', processed_data_origin_file, self.get_maximal_frequent_itemset_sample_number_list_final())
        self.get_max_frequent_association_rules_dataframe()['origin_sample_antecedents'] = self.get_maximal_frequent_itemset_sample_number_list_final()
        self.set_maximal_frequent_itemset_sample_number_list_final([])
        self.get_sample_origin_for_association_rules(self.get_max_frequent_association_rules_dataframe(), 'maximum_frequent_consequents', processed_data_origin_file, self.get_maximal_frequent_itemset_sample_number_list_final())
        self.get_max_frequent_association_rules_dataframe()['origin_sample_consequents'] = self.get_maximal_frequent_itemset_sample_number_list_final()
        self.set_maximal_frequent_itemset_sample_number_list_final([])
        self.get_max_frequent_association_rules_dataframe()['support'] = self.get_maximal_frequent_itemset_supports()
        self.get_max_frequent_association_rules_dataframe()['antecedent support'] = self.get_maximal_frequent_itemset_antecedent_supports()
        self.get_max_frequent_association_rules_dataframe()['consequent support'] = self.get_maximal_frequent_itemset_consequent_supports()
        self.get_max_frequent_association_rules_dataframe()['confidence'] = self.get_maximal_frequent_itemset_confidences()
        self.get_max_frequent_association_rules_dataframe()['lift'] = self.get_maximal_frequent_itemset_lifts()
        self.get_max_frequent_association_rules_dataframe()['leverage'] = self.get_maximal_frequent_itemset_leverages()
        self.get_max_frequent_association_rules_dataframe()['conviction'] = self.get_maximal_frequent_itemset_convictions()
        self.get_max_frequent_association_rules_dataframe()['zhangs_metric'] = self.get_maximal_frequent_itemset_zhangs_metrics()
        self.get_max_frequent_association_rules_dataframe().sort_values(by = ['length_antecedents', 'length_consequents', 'support'], ascending = [False, False, False], inplace = True)
        self.get_max_frequent_association_rules_dataframe().reset_index(drop = True, inplace = True)
        self.set_max_frequent_association_rules_stat_df(self.get_max_frequent_association_rules_dataframe().value_counts(subset = ['length_antecedents', 'length_consequents']))
        self.create_by_sample_dataframe_for_association_rules(self.get_max_frequent_association_rules_dataframe(), 'maximum_frequent_antecedents', 'maximum_frequent_consequents',
							      self.get_maximum_frequent_itemset_association_df_by_sample())        
        self.get_maximum_frequent_itemset_association_df_by_sample().sort_values(by = 'common_sample_number', inplace = True)
        self.get_maximum_frequent_itemset_association_df_by_sample().reset_index(drop = True, inplace = True)
        self.set_maximum_frequent_itemset_association_df_sample_matrix(self.create_association_mining_matrix(self.get_maximum_frequent_itemset_association_df_by_sample(), 'common_sample_number',
								       'antecedent_with_consequents', self.get_maximum_frequent_itemset_association_df_sample_matrix()))

    '''
    Make dataframes empty
    '''
    def make_dataframes_empty(self):
        self.set_frequent_itemset_df(pd.DataFrame(columns = ['frequent_itemsets', 'origin_sample', 'support', 'length']))
        self.set_closed_itemset_df(pd.DataFrame(columns = ['closed_itemsets', 'origin_sample', 'support', 'length']))
        self.set_maximum_frequent_itemset_df(pd.DataFrame(columns = ['maximum_frequent_itemsets', 'origin_sample', 'support', 'length']))
        self.set_association_dataframe(pd.DataFrame())
        self.set_closed_association_rules_dataframe(pd.DataFrame(columns = ['closed_antecedents', 'closed_consequents', 'length_antecedents','length_consequents','origin_sample_antecedents',
									    'origin_sample_consequents', 'antecedent support', 'consequent support', 'support', 'confidence', 'lift', 'leverage',
									    'conviction', 'zhangs_metric']))
        self.set_max_frequent_association_rules_dataframe(pd.DataFrame(columns = ['maximum_frequent_antecedents', 'maximum_frequent_consequents', 'length_antecedents','length_consequents',
										  'origin_sample_antecedents', 'origin_sample_consequents', 'antecedent support', 'consequent support', 'support',
										  'confidence', 'lift', 'leverage', 'conviction', 'zhangs_metric']))
        self.set_frequent_itemset_summary_stat_df(pd.DataFrame())
        self.set_closed_itemset_summary_stat_df(pd.DataFrame())
        self.set_maximum_frequent_itemset_summary_stat_df(pd.DataFrame())
        self.set_frequent_itemset_df_by_item(pd.DataFrame(columns = ['item', 'itemset_length_list', 'origin_sample', 'count']))
        self.set_closed_itemset_df_by_item(pd.DataFrame(columns = ['item', 'itemset_length_list', 'origin_sample', 'count']))
        self.set_maximum_frequent_itemset_df_by_item(pd.DataFrame(columns = ['item', 'itemset_length_list', 'origin_sample', 'count']))
        self.set_frequent_itemset_sample_number_list_final([])
        self.set_closed_itemset_sample_number_list_final([])
        self.set_maximal_frequent_itemset_sample_number_list_final([])
        self.set_association_dataframe_summary_stat_df(pd.DataFrame())
        self.set_closed_association_rules_stat_df(pd.DataFrame())
        self.set_max_frequent_association_rules_stat_df(pd.DataFrame())
        self.set_closed_itemset_df_by_sample(pd.DataFrame())
        self.set_closed_itemset_association_df_by_sample(pd.DataFrame())
        self.set_maximum_frequent_itemset_df_by_sample(pd.DataFrame())
        self.set_maximum_frequent_itemset_association_df_by_sample(pd.DataFrame())
        self.set_closed_itemset_df_sample_matrix(pd.DataFrame())
        self.set_closed_itemset_association_df_sample_matrix(pd.DataFrame())       
        self.set_maximum_frequent_itemset_df_sample_matrix(pd.DataFrame())
        self.set_maximum_frequent_itemset_association_df_sample_matrix(pd.DataFrame())
        

    '''
    Find the origin sample for itemset mining
    '''
    def get_sample_origin(self, itemset_dataframe, input_dataframe, itemset_column, final_list):
        for item in itemset_dataframe[f'{itemset_column}']:
            self.temp_list_2 = []
            for thing in item:
                self.temp_list = []
                for element in input_dataframe.columns:
                    if element == thing:
                        for i, member in enumerate(input_dataframe[f"{element}"]):
                            if member == True:
                                self.temp_list.append(input_dataframe.index.values[i])
                self.temp_list_2.append(self.temp_list)
            final_list.append(self.temp_list_2)

    '''
    Find the origin sample for association rules mining
    '''
    def get_sample_origin_for_association_rules(self, itemset_dataframe, itemset_column, processed_data_origin_file, final_list):
        try:
            self.processed_data_origin_dataframe = pd.read_csv(f'{SAVE_PATH}/{processed_data_origin_file}.csv', low_memory = False, index_col = 0)
        except:
            print('The input file is missing.')
        try:
            self.processed_data_origin_dataframe = self.processed_data_origin_dataframe.loc[:, ~self.processed_data_origin_dataframe.columns.str.contains('^Unnamed')]
        except:
            pass
        for item in itemset_dataframe[f'{itemset_column}']:
            self.temp_list_2 = []
            for thing in item:
                self.temp_list = []
                for element in self.processed_data_origin_dataframe.columns:
                    if element == thing:
                        for i, member in enumerate(self.processed_data_origin_dataframe[f"{element}"]):
                            if member == True:
                                self.temp_list.append(self.processed_data_origin_dataframe.index.values[i])
                self.temp_list_2.append(self.temp_list)
            final_list.append(self.temp_list_2)

    '''
    Create a matrix for itemset visualisation
    '''
    def create_itemset_mining_matrix(self, input_dataframe, sample_numnber_column, itemset_column, output_dataframe):
        self.unique_column_list = []
        self.unique_index_list = []     
        self.unique_index_list = input_dataframe[f"{sample_numnber_column}"].unique()
        self.unique_column_list = input_dataframe[f"{itemset_column}"].unique()

        output_dataframe = pd.DataFrame(columns = self.unique_column_list, index = self.unique_index_list)
        output_dataframe = output_dataframe.convert_dtypes(infer_objects = False, convert_string = False, convert_integer = False, convert_floating = False)

        for i, item in enumerate(input_dataframe[f"{itemset_column}"]):
            output_dataframe.at[input_dataframe.at[i,f"{sample_numnber_column}"], item] = True
        
        for item in output_dataframe.columns:      
            for i, thing in enumerate(output_dataframe[item]):
                if thing != True:
                    output_dataframe.at[output_dataframe.index.values[i], item] = False
        
        return output_dataframe

    '''
    Create a matrix for itemset visualisation
    '''
    def create_association_mining_matrix(self, input_dataframe, sample_numnber_column, itemset_column, output_dataframe):
        self.unique_column_list = []
        self.unique_index_list = []     
        self.unique_index_list = input_dataframe[f"{sample_numnber_column}"].unique()
        self.unique_column_list = input_dataframe[f"{itemset_column}"].unique()

        output_dataframe = pd.DataFrame(columns = self.unique_column_list, index = self.unique_index_list)
        output_dataframe = output_dataframe.convert_dtypes(infer_objects = False, convert_string = False, convert_integer = False, convert_floating = False)

        for i, item in enumerate(input_dataframe[f"{itemset_column}"]):
            output_dataframe.at[input_dataframe.at[i,f"{sample_numnber_column}"], item] = True
        
        for item in output_dataframe.columns:
            for i, thing in enumerate(output_dataframe[item]):
                if thing != True:
                    output_dataframe.at[output_dataframe.index.values[i], item] = False

        return output_dataframe

    '''
    Run the FPGrowth mining and create basic summary statistics of the reusults
    '''
    def FPGrowth(self, abundance_threshold, min_support, input_file):

        self.set_mining_output_df(self.update_values(abundance_threshold, input_file))
        self.set_frequent_itemsets(fpgrowth(self.get_mining_output_df(), min_support = min_support, use_colnames=True))

        self.calculate_closed_itemsets(self.get_frequent_itemsets(), 'support', 'itemsets')
        self.calculate_max_frequent_itemsets(self.get_frequent_itemsets(), 'support', 'itemsets')
        
        self.assign_lists_to_dataframes()
        self.set_closed_and_max_frequent_itemset_lists_to_zero()

        if not os.path.exists(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}.xlsx"):
            workbook = openpyxl.Workbook()
            workbook.create_sheet("All itemsets")
            workbook.create_sheet("All itemsets stat")
            workbook.create_sheet("All itemsets by item")
            workbook.create_sheet("Closed itemsets")
            workbook.create_sheet("Closed itemsets stat")
            workbook.create_sheet("Closed itemsets by item")
            workbook.create_sheet("Closed itemsets by sample")
            workbook.create_sheet("Closed itemsets by sample mtx")
            workbook.create_sheet("Max frequent itemsets")
            workbook.create_sheet("Max frequent itemsets stat")
            workbook.create_sheet("Max frequent itemsets by item")
            workbook.create_sheet("Max frequent itemsets by sample")
            workbook.create_sheet("Max freq sets by sample mtx")
            workbook.save(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}.xlsx")
        
        with pd.ExcelWriter(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}.xlsx", engine="openpyxl", mode = 'a', if_sheet_exists = 'overlay') as writer:
            try:
                self.get_frequent_itemset_df().to_excel(writer, sheet_name="All itemsets", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_frequent_itemset_df().to_csv(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_all_itemsets.csv")
            self.get_frequent_itemset_summary_stat_df().to_excel(writer, sheet_name="All itemsets stat", startrow=0, startcol=0, header=True, index=True)
            self.get_frequent_itemset_df_by_item().to_excel(writer, sheet_name="All itemsets by item", startrow=0, startcol=0, header=True, index=True)
            try:
                self.get_closed_itemset_df().to_excel(writer, sheet_name="Closed itemsets", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_itemset_df().to_csv(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_closed_itemsets.csv")
            self.get_closed_itemset_summary_stat_df().to_excel(writer, sheet_name="Closed itemsets stat", startrow=0, startcol=0, header=True, index=True)
            try:
                self.get_closed_itemset_df_by_item().to_excel(writer, sheet_name="Closed itemsets by item", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_itemset_df_by_item().to_csv(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_closed_itemsets_by_item.csv")
            try:
                self.get_closed_itemset_df_by_sample().to_excel(writer, sheet_name="Closed itemsets by sample", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_itemset_df_by_sample().to_csv(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_closed_itemsets_by_sample.csv")
            try:
                self.get_closed_itemset_df_sample_matrix().to_excel(writer, sheet_name="Closed itemsets by sample mtx", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_itemset_df_sample_matrix().to_csv(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_closed_itemsets_by_sample_mtx.csv")
            try:
                self.get_maximum_frequent_itemset_df().to_excel(writer, sheet_name="Max frequent itemsets", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_maximum_frequent_itemset_df().to_csv(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_max_frequent_itemsets.csv")
            self.get_maximum_frequent_itemset_summary_stat_df().to_excel(writer, sheet_name="Max frequent itemsets stat", startrow=0, startcol=0, header=True, index=True)
            try:
                self.get_maximum_frequent_itemset_df_by_item().to_excel(writer, sheet_name="Max frequent itemsets by item", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_maximum_frequent_itemset_df_by_item().to_csv(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_max_frequent_itemsets_by_item.csv")
            try:
                self.get_maximum_frequent_itemset_df_by_sample().to_excel(writer, sheet_name="Max frequent itemsets by sample", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_maximum_frequent_itemset_df_by_sample().to_csv(f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_max_frequent_itemsets_by_sample.csv")
            try:
                self.get_maximum_frequent_itemset_df_sample_matrix().to_excel(writer, sheet_name="Max freq sets by sample mtx", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_maximum_frequent_itemset_df_sample_matrix().to_csv(\
		f"{SAVE_PATH}//fpgrowth_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_max_frequent_itemsets_by_sample_mtx.csv")
            
            workbook = writer.book
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

        self.make_dataframes_empty()
        
        print("\n Output file is saved")

    '''
    Run the Apriori mining and create basic summary statistics of the results.
    '''
    def Apriori(self, abundance_threshold, min_support, input_file):
        self.set_mining_output_df(self.update_values(abundance_threshold, input_file))
        self.set_frequent_itemsets(fpgrowth(self.get_mining_output_df(), min_support = min_support, use_colnames=True))

        self.calculate_closed_itemsets(self.get_frequent_itemsets(), 'support', 'itemsets')
        self.calculate_max_frequent_itemsets(self.get_frequent_itemsets(), 'support', 'itemsets')
        self.assign_lists_to_dataframes()
        self.set_closed_and_max_frequent_itemset_lists_to_zero()

        if not os.path.exists(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}.xlsx"):
            workbook = openpyxl.Workbook()
            workbook.create_sheet("All itemsets")
            workbook.create_sheet("All itemsets stat")
            workbook.create_sheet("All itemsets by item")
            workbook.create_sheet("Closed itemsets")
            workbook.create_sheet("Closed itemsets stat")
            workbook.create_sheet("Closed itemsets by item")
            workbook.create_sheet("Closed itemsets by sample")
            workbook.create_sheet("Closed itemsets by sample mtx")
            workbook.create_sheet("Max frequent itemsets")
            workbook.create_sheet("Max frequent itemsets stat")
            workbook.create_sheet("Max frequent itemsets by item")
            workbook.create_sheet("Max frequent itemsets by sample")
            workbook.create_sheet("Max freq sets by sample mtx")
            workbook.save(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}.xlsx")

        with pd.ExcelWriter(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}.xlsx", engine="openpyxl", mode = 'a', if_sheet_exists = 'overlay') as writer:
            try:
                self.get_frequent_itemset_df().to_excel(writer, sheet_name="All itemsets", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_frequent_itemset_df().to_csv(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_all_itemsets.csv")
            self.get_frequent_itemset_summary_stat_df().to_excel(writer, sheet_name="All itemsets stat", startrow=0, startcol=0, header=True, index=True)
            try:
                self.get_frequent_itemset_df_by_item().to_excel(writer, sheet_name="All itemsets by item", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_frequent_itemset_df_by_item().to_csv(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_all_itemsets_by_item.csv")
            try:
                self.get_closed_itemset_df().to_excel(writer, sheet_name="Closed itemsets", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_itemset_df().to_csv(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_closed_itemsets.csv")
            self.get_closed_itemset_summary_stat_df().to_excel(writer, sheet_name="Closed itemsets stat", startrow=0, startcol=0, header=True, index=True)
            try:
                self.get_closed_itemset_df_by_item().to_excel(writer, sheet_name="Closed itemsets by item", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_itemset_df_by_item().to_csv(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_closed_itemsets_by_item.csv")
            try:
                self.get_closed_itemset_df_by_sample().to_excel(writer, sheet_name="Closed itemsets by sample", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_itemset_df_by_sample().to_csv(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_closed_itemsets_by_sample.csv")
            try:
                self.get_closed_itemset_df_sample_matrix().to_excel(writer, sheet_name="Closed itemsets by sample mtx", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_itemset_df_sample_matrix().to_csv(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_closed_itemsets_by_sample_mtx.csv")
            try:
                self.get_maximum_frequent_itemset_df().to_excel(writer, sheet_name="Max frequent itemsets", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_maximum_frequent_itemset_df().to_csv(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_max_frequent_itemsets.csv")
            self.get_maximum_frequent_itemset_summary_stat_df().to_excel(writer, sheet_name="Max frequent itemsets stat", startrow=0, startcol=0, header=True, index=True)
            try:
                self.get_maximum_frequent_itemset_df_by_item().to_excel(writer, sheet_name="Max frequent itemsets by item", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_maximum_frequent_itemset_df_by_item().to_csv(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_max_frequent_itemsets_by_item.csv")
            try:
                self.get_maximum_frequent_itemset_df_by_sample().to_excel(writer, sheet_name="Max frequent itemsets by sample", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_maximum_frequent_itemset_df_by_sample().to_csv(f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_max_frequent_itemsets_by_sample.csv")
            try:
                self.get_maximum_frequent_itemset_df_sample_matrix().to_excel(writer, sheet_name="Max freq sets by sample mtx", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_maximum_frequent_itemset_df_sample_matrix().to_csv(\
		f"{SAVE_PATH}//apriori_mining_{str(input_file)[26:]}_abundance_threshold_{abundance_threshold}_min_support_{min_support}_max_frequent_itemsets_by_sample_mtx.csv")

            workbook = writer.book
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

            self.make_dataframes_empty()
        
            print("\n Output file is saved")

    '''
    Update the dataframe for association rules mining.
    '''
    def update_association_dataframes(self, input_file):
        try:
            self.set_input_dataframe(pd.read_excel(f'{SAVE_PATH}/{input_file}.xlsx', sheet_name = 'All itemsets'))
        except:
           print('The input file is missing.')
        try:
            self.set_input_dataframe(self.get_input_dataframe.loc()[:, ~self.input_dataframe.columns.str.contains('^Unnamed')])
        except:
            pass
        
        if 'frequent_itemsets' in self.get_input_dataframe().columns:
            self.get_input_dataframe()['itemsets'] = self.get_input_dataframe()['frequent_itemsets']
            self.get_input_dataframe().drop(['frequent_itemsets'], axis = 1, inplace = True)
        else:
            print("The frequent_itemsets column is missing from the input file")
        
        try:
            self.get_input_dataframe().drop(['length', 'origin_sample'], axis = 1, inplace = True)
        except:
            print("The column titles are not present, the columns can't be dropped")


        for i, item in enumerate(self.get_input_dataframe()['itemsets']):
            item = item.replace("frozenset({", "")
            item = item.replace("})", "")
            item = item.replace(",", "")
            item = item.replace("'", "")
            self.split_item = item.split(' ')
            self.temp_list = []
            for member in self.split_item:
                self.temp_list.append(member)
            self.frozen_list = frozenset(self.temp_list)
            self.get_input_dataframe().at[i, 'itemsets'] = self.frozen_list
            self.frozen_list = []

    '''
    Run the association rules mining.
    '''   
    def Association_rules_mining(self, input_file, metric, metric_threshold, processed_data_origin, processed_data_origin_file):

        self.update_association_dataframes(f"{input_file}")

        try:
            self.set_association_dataframe(association_rules(self.get_input_dataframe(), metric = metric, min_threshold=metric_threshold))
            
        except:
            self.set_association_dataframe(association_rules(self.get_input_dataframe(), support_only=True, min_threshold=metric_threshold))

        self.get_association_dataframe()['length_antecedents'] = self.get_association_dataframe()['antecedents'].apply(lambda x: len(x))
        self.get_association_dataframe()['length_consequents'] = self.get_association_dataframe()['consequents'].apply(lambda x: len(x))
        self.get_sample_origin_for_association_rules(self.get_association_dataframe(), 'antecedents', processed_data_origin_file, self.get_frequent_association_rules_list_final())
        self.get_association_dataframe()['origin_sample_antecedents'] = self.get_frequent_association_rules_list_final()
        self.set_frequent_association_rules_list_final([])
        self.get_sample_origin_for_association_rules(self.get_association_dataframe(), 'consequents', processed_data_origin_file, self.get_frequent_association_rules_list_final())
        self.get_association_dataframe()['origin_sample_consequents'] = self.get_frequent_association_rules_list_final()
        self.set_frequent_association_rules_list_final([])
        self.get_association_dataframe().sort_values(by = ['length_antecedents', 'length_consequents', 'support'], ascending = [False, False, False], inplace = True)
        self.get_association_dataframe().reset_index(drop = True, inplace = True)
        self.set_association_dataframe_summary_stat_df(self.get_association_dataframe().value_counts(subset = ['length_antecedents', 'length_consequents']))

        self.calculate_closed_itemsets_for_association_rules(self.get_association_dataframe(), 'support', 'antecedents')
        self.calculate_max_frequent_itemsets_for_association_rules(self.get_association_dataframe(), 'support', 'antecedents')
        
        self.assign_lists_to_dataframes_for_association_rules(processed_data_origin_file)
        self.set_closed_and_max_frequent_itemset_lists_to_zero()
        
        if not os.path.exists(f"{SAVE_PATH}//association_rules_mining_{processed_data_origin}_metric_{metric}_min_threshold_{metric_threshold}_input_file.xlsx"):
            workbook = openpyxl.Workbook()
            workbook.create_sheet("Association rules")
            workbook.create_sheet("Association rules stat")
            workbook.create_sheet("Closed assoc. rules")
            workbook.create_sheet("Closed assoc. rules stat")
            workbook.create_sheet("Closed assoc. by sample")
            workbook.create_sheet("Closed assoc. by sample mtx")
            workbook.create_sheet("Max freq. assoc. rules")
            workbook.create_sheet("Max freq. assoc. rules stat")
            workbook.create_sheet("Max freq. assoc. by sample")
            workbook.create_sheet("Max freq. assoc. by sample mtx")
            workbook.save(f"{SAVE_PATH}//association_rules_mining_{processed_data_origin}_metric_{metric}_min_threshold_{metric_threshold}_input_file.xlsx")
        
        with pd.ExcelWriter(f"{SAVE_PATH}//association_rules_mining_{processed_data_origin}_metric_{metric}_min_threshold_{metric_threshold}_input_file.xlsx", engine="openpyxl", mode = 'a', if_sheet_exists = 'overlay') as writer:
            try:
                self.get_association_dataframe().to_excel(writer, sheet_name="Association rules", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_association_dataframe().to_csv(f"{SAVE_PATH}//association_rules_mining_{processed_data_origin}_metric_{metric}_min_threshold_{metric_threshold}_association_rules_dataframe.csv")  
            self.get_association_dataframe_summary_stat_df().to_excel(writer, sheet_name="Association rules stat", startrow=0, startcol=0, header=True, index=True)
            try:
                self.get_closed_association_rules_dataframe().to_excel(writer, sheet_name="Closed assoc. rules", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_association_rules_dataframe().to_csv(f"{SAVE_PATH}//association_rules_mining_{processed_data_origin}_metric_{metric}_min_threshold_{metric_threshold}_closed_association_rules_dataframe.csv")
            self.get_closed_association_rules_stat_df().to_excel(writer, sheet_name="Closed assoc. rules stat", startrow=0, startcol=0, header=True, index=True)
            try:
                self.get_closed_itemset_association_df_by_sample().to_excel(writer, sheet_name="Closed assoc. by sample", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_itemset_association_df_by_sample().to_csv(f"{SAVE_PATH}//association_rules_mining_{processed_data_origin}_metric_{metric}_min_threshold_{metric_threshold}_closed_itemset_association_df_by_sample.csv")
            try:
                self.get_closed_itemset_association_df_sample_matrix().to_excel(writer, sheet_name="Closed assoc. by sample mtx", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_closed_itemset_association_df_sample_matrix().to_csv(\
		f"{SAVE_PATH}//association_rules_mining_{processed_data_origin}_metric_{metric}_min_threshold_{metric_threshold}_closed_itemset_association_df_sample_mtx.csv")
            try:
            	self.get_max_frequent_association_rules_dataframe().to_excel(writer, sheet_name="Max freq. assoc. rules", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_max_frequent_association_rules_dataframe().to_csv(f"{SAVE_PATH}//association_rules_mining_{processed_data_origin}_metric_{metric}_min_threshold_{metric_threshold}_max_frequent_association_rules_dataframe.csv")
            self.get_max_frequent_association_rules_stat_df().to_excel(writer, sheet_name="Max freq. assoc. rules stat", startrow=0, startcol=0, header=True, index=True)
            try:
                self.get_maximum_frequent_itemset_association_df_by_sample().to_excel(writer, sheet_name="Max freq. assoc. by sample", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_maximum_frequent_itemset_association_df_by_sample().to_csv(\
		f"{SAVE_PATH}//association_rules_mining_{processed_data_origin}_metric_{metric}_min_threshold_{metric_threshold}_maximum_frequent_itemset_association_df_by_sample.csv")
            try:
                self.get_maximum_frequent_itemset_association_df_sample_matrix().to_excel(writer, sheet_name="Max freq. assoc. by sample mtx", startrow=0, startcol=0, header=True, index=True)
            except:
                self.get_maximum_frequent_itemset_association_df_sample_matrix().to_csv(\
		f"{SAVE_PATH}//association_rules_mining_{processed_data_origin}_metric_{metric}_min_threshold_{metric_threshold}_maximum_frequent_itemset_association_df_sample_matx.csv")
                
            workbook = writer.book
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

        self.make_dataframes_empty()
        
        print("\n Output file is saved, if the output file has only support values are avaialble, then the support only flag was automatically activated.")

mn = Mining()
try:
    input_commands_file = open(f"{SAVE_PATH}//mining_input_commands.txt", "r")
except:
    print("The input_command.txt file is missing")
input_commands_file_content = input_commands_file.readlines()
for line_number, input_commands_file_content_line in enumerate(input_commands_file_content):
    if input_commands_file_content_line[0] == "#":
        pass
    elif input_commands_file_content_line[0] == "\n":
        pass
    elif input_commands_file_content_line[0] == "":
        pass
    elif input_commands_file_content_line[0] == " ":
        print(f"A whitespace is present at line {line_number}, please delete it.")
    else:
        print(f"Command {input_commands_file_content_line} is running.")
        eval(f"mn.{input_commands_file_content_line}")
try:
    input_commands_file.close()
except:
    print("The input_commands_file can't be closed")
